import pandas as pd
from participants import Participants
from groups import Groups
from matches import Matches
from openpyxl import load_workbook

file_path = 'entries.xlsx'

# Read the file as an object and read the entries
xl = pd.ExcelFile(file_path)
df = pd.read_excel(xl)

# Load an existing workbook
wb = load_workbook(file_path)

def is_sheet_name_exist(wb, sheet_name):
    for ws in wb.sheetnames:
        if ws == sheet_name:
            return True
    return False

num_sheets = len(xl.sheet_names)
if num_sheets == 1:
    num_groups = int(input("How many groups?"))
else:
    num_groups = 4

prt = Participants(df)
grp = Groups(num_groups, prt.list_participants)
fxt = Matches()

# TODO: create a dictionary of sheet names and do the check within one line
# TODO: adjust the alignment of the texts on cells
if not is_sheet_name_exist(wb, 'Group Stage'):
    ws1 = wb.create_sheet(title='Group Stage')
    ws1 = grp.print_groups(ws1)

# TODO: Point calculator using winner/loser information
if not is_sheet_name_exist(wb, 'Matches'):
    ws2 = wb.create_sheet(title='Matches')
    matches = fxt.create_matches(grp.groups)
    ws2 = fxt.print_matches(ws2)


wb.save(file_path)



