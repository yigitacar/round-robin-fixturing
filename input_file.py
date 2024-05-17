import pandas as pd
from openpyxl import load_workbook
class InputFile:
    def __init__(self, file_path):
        self.file_path = file_path
        self.num_sheets = 0
        self.entries = self.read_entries()
        self.wb = load_workbook(self.file_path)

    def read_entries(self):
        # Read the file as an object and read the entries
        xl = pd.ExcelFile(self.file_path)
        self.num_sheets = len(xl.sheet_names)
        df = pd.read_excel(xl)
        return df
    def is_sheet_name_exist(self, sheet_name):
        for ws in self.wb.sheetnames:
            if ws == sheet_name:
                return True
        return False

