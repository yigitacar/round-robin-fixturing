from openpyxl.styles import Font
import random
START_LETTER = 'A'
class Matches:
    def __init__(self):
        self.groups = []
        self.num_teams = 0
        self.num_rounds = 0
        self.half_teams = 0
        self.matches = []
        self.match_dictionary = {}

    def initialize(self, groups):
        self.groups = groups
        self.num_teams = len(groups[1])
        self.num_rounds = self.num_teams - 1
        self.half_teams = self.num_teams // 2

    def shuffle_teams(self):
        for group in self.groups:
            random.shuffle(group)

    def create_dictionary(self):
        self.match_dictionary = {
            'matches': [],
        }

        for i, group in enumerate(self.matches):
            match_rounds = {
                'group': chr(ord(START_LETTER) + i),
                'rounds': [],
            }
            for j, match_list in enumerate(group):
                match_round = {
                    'round_num': j + 1,
                    'match_list': match_list,
                    'winners': [],
                    'losers': []
                }
                match_rounds['rounds'].append(match_round)
            self.match_dictionary['matches'].append(match_rounds)
        #print(self.match_dictionary)

    def create_matches(self, groups):
        self.initialize(groups)
        self.shuffle_teams()
        for group in self.groups:
            fixture = self.create_fixture(group)
            self.matches.append(fixture)
        self.create_dictionary()
        return self.matches

    def create_fixture(self, teams):
        fixture = []
        for round_num in range(self.num_rounds):
            round_matches = []
            for i in range(self.half_teams):
                if teams[i] is not None and teams[self.num_teams - i - 1] is not None:
                    round_matches.append((teams[i], teams[self.num_teams - i - 1]))
            fixture.append(round_matches)

            # Rotate the teams for the next round
            teams = [teams[0]] + [teams[-1]] + teams[1:-1]

        return fixture

    def print_matches(self, work_sheet):
        match_dictionary = self.match_dictionary
        for i, group in enumerate(match_dictionary['matches']):
            row_num = i*(self.half_teams+3)+1
            group_cell = work_sheet.cell(row=row_num, column=1)
            group_cell.value = 'Group ' + group['group']
            group_cell.font = Font(size=14, bold=True)
            for j, match_round in enumerate(group['rounds']):
                title_cell = work_sheet.cell(row=row_num+1, column=j+1)
                title_cell.value = 'Round ' + str(match_round['round_num'])
                title_cell.font = Font(size=12, bold=True)
                for k, match in enumerate(match_round['match_list']):
                    entry = match[0] + ' vs ' + match[1]
                    work_sheet.cell(row=row_num+k+2, column=j+1).value = entry

        return work_sheet

