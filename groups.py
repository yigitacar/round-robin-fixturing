from openpyxl.styles import Font, PatternFill
START_LETTER = 'A'
POINT_COLUMN = 12
class Groups:
    def __init__(self, num_groups, participants):
        self.num_groups = num_groups
        self.num_participants = len(participants)
        self.group_size = self.num_participants // self.num_groups
        self.groups = []
        self.distribute_into_groups(participants)
        self.group_dictionary = {}
        self.create_dictionary()

    def distribute_into_groups(self, participants):
        self.groups = [participants[i * self.group_size:(i + 1) * self.group_size] for i in range(self.num_groups - 1)]
        self.groups.append(participants[(self.num_groups - 1) * self.group_size:])
    def is_group_num_valid(self):
        return self.num_participants % self.num_groups == 0

    def create_dictionary(self):
        self.group_dictionary = {
            'groups': [],
        }

        for ind, participants in enumerate(self.groups):
            group = {
                'num': chr(ord(START_LETTER) + ind),
                'participants': participants,
                'points': [0] * self.group_size
            }
            self.group_dictionary['groups'].append(group)
    def get_groups(self):
        return self.groups

    def print_groups(self, work_sheet):
        for i, group in enumerate(self.group_dictionary['groups']):
            row_num = i*(self.group_size+2)+1
            group_cell = work_sheet.cell(row=row_num, column=1)
            group_cell.value = 'Group ' + group['num']
            group_cell.font = Font(size=14, bold=True)
            title_cell1 = work_sheet.cell(row=row_num, column=11)
            title_cell1.value = 'Participants'
            title_cell1.font = Font(bold=True)
            title_cell2 = work_sheet.cell(row=row_num, column=12)
            title_cell2.value = 'Points'
            title_cell2.font = Font(bold=True)
            bye_cell_row_num = row_num + 1
            bye_cell_column_num = 2

            for j, entry in enumerate(group['participants']):
                work_sheet.cell(row=row_num+j+1, column=1).value = entry
                work_sheet.cell(row=row_num, column=j+2).value = entry
                work_sheet.cell(row=row_num+j+1, column=11).value = entry
                bye_cell = work_sheet.cell(row=bye_cell_row_num, column=bye_cell_column_num)
                bye_cell.fill = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")
                bye_cell_row_num = bye_cell_row_num + 1
                bye_cell_column_num = bye_cell_column_num + 1

        return work_sheet

    def read_groups(self, work_sheet):
        for i, group in enumerate(self.group_dictionary['groups']):
            start_row = i * (self.group_size + 2) + 2
            end_row = start_row + self.group_size - 1
            element_counter = 0
            for j in range(start_row, end_row + 1):
                group['participants'][element_counter] = work_sheet.cell(row=j, column=1).value
                element_counter = element_counter + 1
            self.group_dictionary['groups'][i] = group
    def read_winners(self, work_sheet):
        for i, group in enumerate(self.group_dictionary['groups']):
            group_row = i * (self.group_size + 2) + 1
            start_row = i * (self.group_size + 2) + 2
            end_row = start_row + self.group_size - 1
            start_column = 3
            current_column = start_column
            end_column = self.group_size + 1
            for j in range(start_row, end_row + 1):
                for k in range(current_column, end_column + 1):
                    competitor1 = work_sheet.cell(row=j, column=1).value
                    competitor2 = work_sheet.cell(row=group_row, column=k).value
                    competitor1_index = group['participants'].index(competitor1)
                    competitor2_index = group['participants'].index(competitor2)
                    result = work_sheet.cell(row=j, column=k).value
                    if result == 'Draw':
                        group['points'][competitor1_index] += 1
                        group['points'][competitor2_index] += 1
                    elif result == competitor1:
                        group['points'][competitor1_index] += 2
                    elif result == competitor2:
                        group['points'][competitor2_index] += 2
                current_column = current_column + 1
            #print(group)
    # TODO print titles Participants and Points
    def print_points(self, work_sheet):
        for i, group in enumerate(self.group_dictionary['groups']):
            start_row = i * (self.group_size + 2) + 2
            sorted_participants, sorted_points = self.sort_participants(group)
            for j, participant in enumerate(sorted_participants):
                work_sheet.cell(row=start_row+j, column=POINT_COLUMN-1).value = participant
                work_sheet.cell(row=start_row+j, column=POINT_COLUMN).value = sorted_points[j]

    def sort_participants(self, group):
        part_list = group['participants']
        point_list = group['points']
        participants_points = list(zip(part_list, point_list))
        sorted_participants_points = sorted(participants_points, key=lambda x: x[1], reverse=True)
        sorted_participants, sorted_points = zip(*sorted_participants_points)
        sorted_participants = list(sorted_participants)
        sorted_points = list(sorted_points)
        return sorted_participants, sorted_points
